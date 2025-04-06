#!/usr/bin/env python3
"""
A single Python script demonstrating how to:

1) Create some dummy Excel files (for demo purposes).
2) Read each Excel file into a nested dictionary of the form:
   {
       "SheetName": {
           "A1": {
               "value": <evaluated_value>,
               "formula": <formula_string_if_any_or_empty>
           },
           "A2": { ... },
           ...
       },
       ...
   }
3) Compare two such dictionaries to see differences in cell values/formulas.
4) Print out the comparison result.
"""

import openpyxl
import os


def read_excel_to_dict(file_path):
    """
    Read an Excel file and return a nested dictionary of:
        {
            "SheetName": {
                "A1": {"value": <cell_value>, "formula": <cell_formula_or_empty>},
                "A2": {"value": <cell_value>, "formula": <cell_formula_or_empty>},
                ...
            },
            ...
        }

    We use two workbooks:
      - data_only=True:  to extract the last computed (cached) values.
      - data_only=False: to extract the actual formulas.
    Then we combine them to get both the formula and the computed value.
    """
    # Load workbooks in two modes to get both formulas and their computed values
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
    wb_values = openpyxl.load_workbook(file_path, data_only=True)

    excel_dict = {}

    # Both workbooks should have the same sheet names (assuming no differences).
    # We'll iterate over wb_formulas because it definitely has all the sheets.
    for sheet_name in wb_formulas.sheetnames:
        sheet_formula = wb_formulas[sheet_name]
        sheet_value = wb_values[sheet_name]

        sheet_data = {}
        # Iterate over actual used cells (sheet_formula has the same dimension as sheet_value)
        for row in sheet_formula.iter_rows():
            for cell_f in row:
                cell_v = sheet_value[cell_f.coordinate]
                # If the cell was a formula, openpyxl stores that formula in cell_f.value 
                # (when data_only=False). If not, it's just the cell's literal value.
                # The "data_type == 'f'" check helps to see if it's a formula.
                # However, we can also check if 'cell_f.value' starts with '=' if you want,
                # but 'data_type' is more direct.

                if cell_f.data_type == 'f':
                    # cell_f.value holds the formula string (e.g., "=SUM(A1:A2)")
                    formula = cell_f.value
                else:
                    formula = ""

                # cell_v.value is the computed value (the cached value if there's a formula),
                # or the literal value if there's no formula.
                computed_value = cell_v.value

                sheet_data[cell_f.coordinate] = {
                    "value": computed_value,
                    "formula": formula
                }
        excel_dict[sheet_name] = sheet_data

    return excel_dict


def compare_excel_dicts(dict1, dict2):
    """
    Compare two Excel dictionaries produced by `read_excel_to_dict`.
    Return a dictionary describing the differences.

    The returned dictionary can have a structure like:
    {
       "Sheet1": {
          "A1": {
             "value": ("val_in_dict1", "val_in_dict2")    # only if they differ
             "formula": ("form_in_dict1", "form_in_dict2")# only if they differ
          },
          "A2": {...},
          ...
       },
       "Sheet2": ...
    }
    Only cells that differ in "value" or "formula" will appear in the result.
    """
    differences = {}

    # Collect all sheet names from both dicts
    all_sheets = set(dict1.keys()).union(dict2.keys())

    for sheet in all_sheets:
        sheet_diff = {}
        sheet1_data = dict1.get(sheet, {})
        sheet2_data = dict2.get(sheet, {})

        # Collect all cell coords from both sheets
        all_cells = set(sheet1_data.keys()).union(sheet2_data.keys())
        for cell_coord in all_cells:
            cell1 = sheet1_data.get(cell_coord, {})
            cell2 = sheet2_data.get(cell_coord, {})

            val1 = cell1.get("value", None)
            val2 = cell2.get("value", None)
            formula1 = cell1.get("formula", "")
            formula2 = cell2.get("formula", "")

            cell_subdiff = {}

            # Check value differences
            if val1 != val2:
                cell_subdiff["value"] = (val1, val2)

            # Check formula differences
            if formula1 != formula2:
                cell_subdiff["formula"] = (formula1, formula2)

            # If anything differs, record it
            if cell_subdiff:
                sheet_diff[cell_coord] = cell_subdiff

        # If this sheet has differences, record them
        if sheet_diff:
            differences[sheet] = sheet_diff

    return differences


def create_dummy_excel_files():
    """
    Create two dummy Excel files to demonstrate usage.
    We'll create:
      - file1.xlsx: with some values and formulas
      - file2.xlsx: with slightly altered values/formulas
    """
    # --------------- File 1 ---------------
    wb1 = openpyxl.Workbook()
    ws1_1 = wb1.active  # First sheet
    ws1_1.title = "Sheet1"

    # Populate some cells
    ws1_1["A1"] = 10
    ws1_1["A2"] = 20
    ws1_1["A3"] = "=SUM(A1, A2)"  # Formula

    ws1_1["B1"] = "Hello"
    ws1_1["B2"] = "=CONCAT(B1, \" World\")"  # Another formula

    # Create a second sheet
    ws1_2 = wb1.create_sheet("Sheet2")
    ws1_2["C10"] = 99
    ws1_2["C11"] = "=C10 * 2"

    wb1.save("file1.xlsx")

    # --------------- File 2 ---------------
    wb2 = openpyxl.Workbook()
    ws2_1 = wb2.active
    ws2_1.title = "Sheet1"

    # Slightly different data
    ws2_1["A1"] = 10
    ws2_1["A2"] = 21   # changed
    ws2_1["A3"] = "=SUM(A1, A2)"  # formula is the same

    ws2_1["B1"] = "Hello"
    ws2_1["B2"] = "=CONCAT(B1, \" Universe\")"  # changed from "World" to "Universe"

    # Second sheet
    ws2_2 = wb2.create_sheet("Sheet2")
    ws2_2["C10"] = 99
    ws2_2["C11"] = "=C10 * 3"  # changed multiplier

    wb2.save("file2.xlsx")


def main():
    # 1) Create dummy Excel files for demonstration:
    create_dummy_excel_files()
    print("[INFO] Created 'file1.xlsx' and 'file2.xlsx' for demonstration.")

    # 2) Read each file into dictionaries
    dict1 = read_excel_to_dict("file1.xlsx")
    dict2 = read_excel_to_dict("file2.xlsx")

    # 3) Print them out (just to show what they look like)
    print("\n[INFO] Dictionary for file1.xlsx:")
    print(dict1)
    print("\n[INFO] Dictionary for file2.xlsx:")
    print(dict2)

    # 4) Compare the two dictionaries
    diff = compare_excel_dicts(dict1, dict2)

    # 5) Print the differences
    print("\n[INFO] Differences between file1.xlsx and file2.xlsx:")
    print(diff)

    # Optionally, clean up the files (comment out if you want to keep them)
    # os.remove("file1.xlsx")
    # os.remove("file2.xlsx")


if __name__ == "__main__":
    main()
