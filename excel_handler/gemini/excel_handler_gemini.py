import openpyxl
import os
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime
import pytz # Import pytz for timezone handling
import argparse # Import argparse for command-line arguments
import time # Import time for potential retries

# --- Core Data Extraction Functions (Unchanged) ---

def get_cell_data(cell_value, cell_formula):
    formula_str = ""
    if isinstance(cell_formula, str) and cell_formula.startswith('='):
        formula_str = cell_formula
    display_value = cell_value if cell_value is not None else '[empty]'
    return {"value": display_value, "formula": formula_str}

def read_worksheet_data(ws_values, ws_formulas):
    sheet_data = {}
    max_row = ws_formulas.max_row
    max_col = ws_formulas.max_column
    if max_row is None or max_col is None or max_row == 0 or max_col == 0:
         return {}

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell_coord = f"{get_column_letter(col_idx)}{row_idx}"
            cell_formulas_obj = ws_formulas.cell(row=row_idx, column=col_idx)
            formula_content = cell_formulas_obj.value
            cell_values_obj = ws_values.cell(row=row_idx, column=col_idx)
            evaluated_value = cell_values_obj.value

            if evaluated_value is not None or (isinstance(formula_content, str) and formula_content.startswith('=')):
                 cell_info = get_cell_data(evaluated_value, formula_content)
                 sheet_data[cell_coord] = cell_info

    return sheet_data

def read_excel_file_data(filepath):
    wb_values = None
    wb_formulas = None
    try:
        retries = 3
        delay = 0.5 # seconds
        last_error = None
        for i in range(retries):
            try:
                wb_values = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
                wb_formulas = openpyxl.load_workbook(filepath, data_only=False, read_only=True)
                last_error = None
                break
            except Exception as e_inner:
                last_error = e_inner
                if wb_values: wb_values.close()
                if wb_formulas: wb_formulas.close()
                wb_values, wb_formulas = None, None
                if i < retries - 1:
                    print(f"  Retrying read for {filepath} after error: {e_inner}")
                    time.sleep(delay)
                else:
                    print(f"  Failed to read {filepath} after {retries} attempts.")

        if wb_values is None or wb_formulas is None:
             if last_error:
                  raise last_error
             else:
                  raise Exception(f"Failed to load workbook '{filepath}' for unknown reasons after retries.")


        all_excel_data = {}
        sheet_names = wb_formulas.sheetnames
        if not sheet_names:
            print(f"  Warning: No sheets found in '{filepath}'.")
            return {}

        for sheet_name in sheet_names:
            if sheet_name in wb_values.sheetnames:
                ws_values = wb_values[sheet_name]
                ws_formulas = wb_formulas[sheet_name]
                print(f"  Processing sheet: '{sheet_name}'...")
                sheet_content = read_worksheet_data(ws_values, ws_formulas)
                if sheet_content:
                     all_excel_data[sheet_name] = sheet_content
                else:
                     print(f"  Sheet '{sheet_name}' appears empty or contains no data/formulas. Adding as empty.")
                     all_excel_data[sheet_name] = {}
            else:
                 print(f"  Warning: Sheet '{sheet_name}' found in formula view but not value view. Skipping.")

        return all_excel_data

    except FileNotFoundError:
        print(f"Error: File not found at '{filepath}'")
        return None
    except Exception as e:
        print(f"Error reading Excel file '{filepath}': {e}")
        return None
    finally:
        if wb_values:
            wb_values.close()
        if wb_formulas:
            wb_formulas.close()

# --- Comparison Function (Unchanged) ---

def compare_excel_data(data1, data2):
    if data1 is None or data2 is None:
        print("Error: Cannot compare None data.")
        return {"_metadata": {"error": "Input data missing"}}

    differences = {}
    sheets1 = set(data1.keys())
    sheets2 = set(data2.keys())

    sheets_only_in_file1 = list(sheets1 - sheets2)
    sheets_only_in_file2 = list(sheets2 - sheets1)
    common_sheets = list(sheets1 & sheets2)

    meta = {
        "sheets_common": sorted(common_sheets),
        "sheets_only_in_file1": sorted(sheets_only_in_file1),
        "sheets_only_in_file2": sorted(sheets_only_in_file2)
    }
    # Add metadata only if there are structural differences or if comparison occurs
    if meta["sheets_only_in_file1"] or meta["sheets_only_in_file2"] or common_sheets:
         differences["_metadata"] = meta # Ensure metadata is always present if sheets exist

    for sheet_name in sorted(common_sheets):
        sheet_diff = {}
        sheet_data1 = data1.get(sheet_name, {})
        sheet_data2 = data2.get(sheet_name, {})
        all_cells = set(sheet_data1.keys()) | set(sheet_data2.keys())

        for cell_coord in sorted(list(all_cells)):
            default_cell = {"value": "[missing]", "formula": ""}
            cell_info1 = sheet_data1.get(cell_coord, default_cell)
            cell_info2 = sheet_data2.get(cell_coord, default_cell)

            values_differ = str(cell_info1["value"]) != str(cell_info2["value"])
            formulas_differ = cell_info1["formula"] != cell_info2["formula"]

            if values_differ or formulas_differ:
                sheet_diff[cell_coord] = {
                    "file1": sheet_data1.get(cell_coord, default_cell),
                    "file2": sheet_data2.get(cell_coord, default_cell)
                }

        if sheet_diff:
            differences[sheet_name] = sheet_diff

    # Ensure metadata is included even if no cell differences found in common sheets
    # (This was implicitly handled before, but making it explicit)
    if not differences and meta:
         differences["_metadata"] = meta

    return differences


# --- Demo File Creation (Unchanged) ---

def create_demo_excel_file(filepath, file_data):
    wb = None
    try:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for sheet_name, sheet_contents in file_data.items():
            ws = wb.create_sheet(title=sheet_name)
            if not sheet_contents:
                 print(f"  Creating empty sheet '{sheet_name}' in '{filepath}'")
                 continue
            for cell_coord, content in sheet_contents.items():
                value_to_write = content
                if isinstance(content, dict):
                     if 'formula' in content and content['formula']:
                          value_to_write = content['formula']
                     elif 'value' in content:
                          value_to_write = content['value'] if content['value'] != '[empty]' else None
                     else:
                          value_to_write = None
                elif content == '[empty]':
                     value_to_write = None
                ws[cell_coord] = value_to_write

        wb.save(filepath)
        print(f"Successfully created demo file: '{filepath}'")
        return True # Indicate success
    except Exception as e:
        print(f"Error creating demo file '{filepath}': {e}")
        return False # Indicate failure
    finally:
        if wb:
            wb.close()


# --- Text File Output Functions (Unchanged) ---

def write_excel_data_to_txt(source_filepath, excel_data, output_filepath):
    print(f"  Writing content dump to '{output_filepath}'...")
    try:
        os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
        with open(output_filepath, 'w', encoding='utf-8') as f:
            f.write(f"Content Dump for Excel File: {source_filepath}\n")
            f.write(f"Generated on (UTC): {datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M:%S %Z')}\n")
            f.write("=" * 40 + "\n\n")
            if not excel_data:
                f.write("No sheets or data found in the file.\n")
                return
            sheet_names = sorted(excel_data.keys())
            f.write(f"Sheets found ({len(sheet_names)}): {', '.join(sheet_names)}\n\n")
            for sheet_name in sheet_names:
                f.write(f"--- Sheet: {sheet_name} ---\n")
                sheet_content = excel_data.get(sheet_name, {})
                if not sheet_content:
                    f.write("  [Sheet is empty or contains no tracked data]\n\n")
                    continue
                # Sort cells based on row number, then column letter
                sorted_cells = sorted(sheet_content.keys(), key=lambda x: (int(''.join(filter(str.isdigit, x))), ''.join(filter(str.isalpha, x))))
                for cell_coord in sorted_cells:
                    cell_info = sheet_content[cell_coord]
                    value_str = cell_info.get('value', '[error retrieving value]')
                    formula_str = cell_info.get('formula', '')
                    f.write(f"  {cell_coord:<8}: Value = {value_str}")
                    if formula_str: f.write(f", Formula = {formula_str}")
                    f.write("\n")
                f.write("\n")
        print(f"  Successfully wrote content dump to '{output_filepath}'")
        return True # Indicate success
    except Exception as e:
        print(f"Error writing content dump to '{output_filepath}': {e}")
        return False # Indicate failure

def write_comparison_summary_to_txt(file1_path, file2_path, comparison_results, output_filepath):
    print(f"  Writing comparison summary to '{output_filepath}'...")
    try:
        os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
        with open(output_filepath, 'w', encoding='utf-8') as f:
            f.write("Excel File Comparison Summary\n")
            f.write(f"Generated on (UTC): {datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M:%S %Z')}\n")
            f.write(f"File 1: {file1_path}\n")
            f.write(f"File 2: {file2_path}\n")
            f.write("=" * 40 + "\n\n")
            if not comparison_results:
                 # Handle case where comparison function itself returned None or empty
                 if comparison_results is None:
                      f.write("Comparison failed or produced no results dictionary.\n")
                 else: # Empty dictionary
                      f.write("Comparison resulted in an empty dictionary (potentially indicating an issue).\n")
                 return False

            meta = comparison_results.get("_metadata")
            # Check if the only content is metadata indicating identical files
            is_identical = False
            if meta and len(comparison_results) == 1: # Only metadata key exists
                 if not meta.get("sheets_only_in_file1") and not meta.get("sheets_only_in_file2"):
                      # Check if there are common sheets listed, implying comparison happened
                      if meta.get("sheets_common"):
                           is_identical = True

            if is_identical:
                 f.write("--- I. Sheet Structure Overview ---\n")
                 common = meta.get("sheets_common", [])
                 f.write(f"Sheets with the same name in BOTH files ({len(common)}): {', '.join(common)}\n")
                 f.write("No sheets found only in File 1.\n")
                 f.write("No sheets found only in File 2.\n\n")
                 f.write("--- II. Detailed Cell Differences (in Common Sheets) ---\n")
                 f.write("No differences found in cell values or formulas within common sheets.\n")
                 return True # Successful summary, no differences found


            # --- Proceed with standard reporting if not identical ---
            if not meta:
                 f.write("Comparison metadata (_metadata) is missing.\n")
                 # Attempt to infer common sheets if differences exist otherwise
                 diff_sheets_keys = [k for k in comparison_results.keys() if k != "_metadata"]
                 if diff_sheets_keys:
                      f.write(f"Differences found in inferred sheets: {', '.join(sorted(diff_sheets_keys))}\n\n")
                 else:
                      f.write("No differences found and metadata is missing.\n")
                 # Cannot reliably report structure overview without metadata
            else:
                 f.write("--- I. Sheet Structure Overview ---\n")
                 common = meta.get("sheets_common", [])
                 only1 = meta.get("sheets_only_in_file1", [])
                 only2 = meta.get("sheets_only_in_file2", [])
                 if common: f.write(f"Sheets with the same name in BOTH files ({len(common)}): {', '.join(common)}\n")
                 else: f.write("No sheets found with the same name in both files.\n")
                 if only1: f.write(f"Sheets found ONLY in File 1 ({len(only1)}): {', '.join(only1)}\n")
                 else: f.write("No sheets found only in File 1.\n")
                 if only2: f.write(f"Sheets found ONLY in File 2 ({len(only2)}): {', '.join(only2)}\n")
                 else: f.write("No sheets found only in File 2.\n")
                 f.write("\n")


            f.write("--- II. Detailed Cell Differences (in Common Sheets) ---\n")
            # Filter out metadata key to get sheets with actual differences
            diff_sheets = sorted([sheet for sheet in comparison_results if sheet != "_metadata"])

            if not diff_sheets:
                 # This case should be covered by the 'is_identical' check now,
                 # but leave a fallback message just in case.
                 f.write("No differences found in cell values or formulas within common sheets.\n")
            else:
                for sheet_name in diff_sheets:
                    sheet_diff_data = comparison_results[sheet_name]
                    # Check if sheet_diff_data is actually a dictionary (it should be)
                    if not isinstance(sheet_diff_data, dict):
                         f.write(f"\n--- Error: Invalid difference data found for Sheet: {sheet_name} ---\n")
                         continue

                    f.write(f"\n--- Differences in Sheet: {sheet_name} ---\n")
                    # Sort differing cells based on row number, then column letter
                    sorted_cells = sorted(sheet_diff_data.keys(), key=lambda x: (int(''.join(filter(str.isdigit, x))), ''.join(filter(str.isalpha, x))))
                    for cell_coord in sorted_cells:
                        diff_info = sheet_diff_data[cell_coord]
                        # Check if diff_info is the expected format
                        if not isinstance(diff_info, dict) or 'file1' not in diff_info or 'file2' not in diff_info:
                             f.write(f"  Cell: {cell_coord} - Error: Malformed difference information.\n\n")
                             continue

                        info1 = diff_info.get('file1', {'value': '[error]', 'formula': ''})
                        info2 = diff_info.get('file2', {'value': '[error]', 'formula': ''})
                        f.write(f"  Cell: {cell_coord}\n")
                        f.write(f"    File 1: Value = {info1.get('value','[N/A]')}")
                        if info1.get('formula'): f.write(f", Formula = {info1['formula']}")
                        f.write("\n")
                        f.write(f"    File 2: Value = {info2.get('value','[N/A]')}")
                        if info2.get('formula'): f.write(f", Formula = {info2['formula']}")
                        f.write("\n\n")
        print(f"  Successfully wrote comparison summary to '{output_filepath}'")
        return True # Indicate success
    except Exception as e:
        print(f"Error writing comparison summary to '{output_filepath}': {e}")
        return False # Indicate failure


# --- NEW: End-to-End Comparison Function ---

def compare_excel_files_e2e(file1_path, file2_path, keep_files=True):
    """
    Performs the end-to-end comparison of two Excel files.

    Reads both files, generates content dump text files, compares data,
    generates a comparison summary text file, and optionally cleans up
    the generated text files and output directory.

    Args:
        file1_path (str): Path to the first Excel file.
        file2_path (str): Path to the second Excel file.
        keep_files (bool, optional): If True, generated text reports and the
                                     output directory are not deleted.
                                     Defaults to False (cleanup occurs).

    Returns:
        dict or None: The comparison results dictionary, or None if a critical
                      error occurred during reading or comparison setup.
                      The dictionary structure details differences found.
                      Returns an empty dict {} if reading failed but allowed continuation.
    """
    print(f"\n--- Starting End-to-End Comparison ---")
    print(f"File 1: {file1_path}")
    print(f"File 2: {file2_path}")
    print(f"Keep generated report files: {keep_files}")

    # --- Define Base Filenames for Reports ---
    # Extract base names from input paths for report naming consistency
    f1_basename = os.path.basename(file1_path)
    f2_basename = os.path.basename(file2_path)
    file1_txt_dump_base = f"{f1_basename}_contents.txt"
    file2_txt_dump_base = f"{f2_basename}_contents.txt"
    comparison_summary_txt_base = f"comparison_{f1_basename}_vs_{f2_basename}.txt"

    # --- Create Timestamped Output Directory ---
    output_dir_name = "." # Default to current dir if error occurs
    file1_txt_dump_path = file1_txt_dump_base
    file2_txt_dump_path = file2_txt_dump_base
    comparison_summary_txt_path = comparison_summary_txt_base
    output_directory_created = False
    try:
        jakarta_tz = pytz.timezone('Asia/Jakarta')
        now_utc = datetime.datetime.now(pytz.utc)
        now_jakarta = now_utc.astimezone(jakarta_tz)
        timestamp_str = now_jakarta.strftime('%Y%m%d_%H%M%S')
        # Make directory name slightly more descriptive
        output_dir_name = f"comparison_output_{timestamp_str}"
        os.makedirs(output_dir_name, exist_ok=True)
        print(f"Created output directory for reports: '{output_dir_name}'")
        output_directory_created = True

        file1_txt_dump_path = os.path.join(output_dir_name, file1_txt_dump_base)
        file2_txt_dump_path = os.path.join(output_dir_name, file2_txt_dump_base)
        comparison_summary_txt_path = os.path.join(output_dir_name, comparison_summary_txt_base)

    except pytz.UnknownTimeZoneError:
         print("Error: Timezone 'Asia/Jakarta' not found. Check pytz installation.")
         print("Text report outputs will be saved in the current directory instead.")
         output_dir_name = "."
    except Exception as e:
        print(f"Error creating output directory or paths: {e}")
        print("Text report outputs will be saved in the current directory instead.")
        output_dir_name = "."

    # --- Read Data from Files ---
    print(f"\n--- Reading Data ---")
    excel1_contents = read_excel_file_data(file1_path)
    excel2_contents = read_excel_file_data(file2_path)

    # Handle cases where reading failed critically
    if excel1_contents is None or excel2_contents is None:
        print("Error: Failed to read one or both Excel files. Comparison cannot proceed.")
        # No comparison results to return in this critical failure case
        return None

    # --- Write Individual Content Dumps ---
    print("\n--- Writing Content Dumps ---")
    write_excel_data_to_txt(file1_path, excel1_contents, file1_txt_dump_path)
    write_excel_data_to_txt(file2_path, excel2_contents, file2_txt_dump_path)

    # --- Compare the Data ---
    print("\n--- Comparing Files ---")
    comparison_results = compare_excel_data(excel1_contents, excel2_contents)

    # --- Write Comparison Summary Report ---
    print("\n--- Writing Comparison Summary ---")
    write_comparison_summary_to_txt(file1_path, file2_path, comparison_results, comparison_summary_txt_path)

    # --- Cleanup Generated Output Files (Conditional) ---
    # Note: This function only cleans up the files IT generated (reports).
    # It does NOT clean up the input Excel files.
    if not keep_files:
        print("\n--- Cleaning Up Generated Report Files ---")
        files_to_clean = [
            file1_txt_dump_path,
            file2_txt_dump_path,
            comparison_summary_txt_path
        ]
        for f_path in files_to_clean:
            try:
                if os.path.exists(f_path) and os.path.isfile(f_path): # Check it's a file
                    os.remove(f_path)
                    print(f"Removed report file: '{f_path}'")
            except Exception as e:
                print(f"Error during cleanup of report file '{f_path}': {e}")

        # Optionally remove the output directory if it was created and is now empty
        try:
            if output_directory_created and os.path.exists(output_dir_name) and os.path.isdir(output_dir_name):
                 if not os.listdir(output_dir_name):
                      os.rmdir(output_dir_name)
                      print(f"Removed empty output directory '{output_dir_name}'")
                 else:
                      print(f"Output directory '{output_dir_name}' not empty after report cleanup, not removed.")
        except Exception as e:
            print(f"Error removing output directory '{output_dir_name}': {e}")
    else:
        print("\n--- Skipping Cleanup of Generated Report Files (as requested) ---")

    print(f"\n--- End-to-End Comparison Finished ---")

    return comparison_results


# --- Main Execution & Demonstration ---

def main():
    # --- Argument Parsing ---
    parser = argparse.ArgumentParser(
        description="Compare two Excel files cell by cell. Generates demo files if run directly, then compares them and reports differences.",
        formatter_class=argparse.RawTextHelpFormatter # Preserve newline formatting in help
        )
    parser.add_argument(
        "--file1",
        type=str,
        help="Path to the first Excel file (optional, uses demo file if not provided)."
    )
    parser.add_argument(
        "--file2",
        type=str,
        help="Path to the second Excel file (optional, uses demo file if not provided)."
    )
    parser.add_argument(
        "--keep-files",
        action="store_true",
        help="If set, prevents the deletion of generated demo Excel files (if created)\nand all generated text report files/directory."
    )
    args = parser.parse_args()

    print("="*50)
    print(" Excel Comparison Utility")
    print("="*50)
    if args.keep_files:
        print(" [--keep-files flag set: Generated files will NOT be deleted.]")


    # --- Determine Input Files (Use demo or provided) ---
    use_demo_files = not (args.file1 and args.file2)
    input_file1 = args.file1
    input_file2 = args.file2
    demo_file1_created = False
    demo_file2_created = False

    if use_demo_files:
        print("\n--- Running in Demo Mode ---")
        print("Input files not specified via --file1 and --file2.")
        print("Generating demo Excel files for comparison.")

        # --- Define Demo Filenames and Data ---
        demo_file1_name = "demo_excel_1.xlsx"
        demo_file2_name = "demo_excel_2.xlsx"
        input_file1 = demo_file1_name
        input_file2 = demo_file2_name

        file1_data_structure = {
            "DataSheet": {
                "A1": {"value": "Name", "formula": ""}, "B1": {"value": "Value1", "formula": ""}, "C1": {"value": "Value2", "formula": ""}, "D1": {"value": "Total", "formula": ""},
                "A2": {"value": "Item A", "formula": ""}, "B2": {"value": 10, "formula": ""}, "C2": {"value": 20, "formula": ""}, "D2": {"value": 30, "formula": "=B2+C2"},
                "A3": {"value": "Item B", "formula": ""}, "B3": {"value": 15, "formula": ""}, "C3": {"value": 25, "formula": ""}, "D3": {"value": 40, "formula": "=SUM(B3:C3)"},
                "E5": {"value": "[empty]", "formula":""}
            },
            "Summary": {
                "A1": {"value": "Grand Total", "formula": ""}, "B1": {"value": 70, "formula": "=DataSheet!D2+DataSheet!D3"}, "C5": {"value": "Only in File 1 Sheet", "formula": ""}
            },
            "IdenticalSheet": {
                 "A1": {"value": "Status", "formula": ""}, "B1": {"value": "OK", "formula": ""}
            },
            "EmptySheet1": {}
        }
        file2_data_structure = {
            "DataSheet": {
                "A1": {"value": "Name", "formula": ""}, "B1": {"value": "Value One", "formula": ""}, "C1": {"value": "Value2", "formula": ""}, "D1": {"value": "Total Sum", "formula": ""},
                "A2": {"value": "Item A", "formula": ""}, "B2": {"value": 10.0, "formula": ""}, "C2": {"value": 20, "formula": ""}, "D2": {"value": 30, "formula": "=B2+C2"},
                "A3": {"value": "Item B New", "formula": ""}, "B3": {"value": 15, "formula": ""}, "C3": {"value": 99, "formula": ""}, "D3": {"value": 114, "formula": "=B3+C3"},
                "F6": {"value": "Extra Cell F6", "formula": ""}
            },
            "Details": {
                "X1": {"value": "Detail Info", "formula": ""}, "Y10": {"value": 12345, "formula": ""}
            },
            "IdenticalSheet": {
                 "A1": {"value": "Status", "formula": ""}, "B1": {"value": "OK", "formula": ""}
            }
        }

        # --- Create Demo Files ---
        print("\n--- Creating Demo Files ---")
        demo_file1_created = create_demo_excel_file(demo_file1_name, file1_data_structure)
        demo_file2_created = create_demo_excel_file(demo_file2_name, file2_data_structure)

        if not (demo_file1_created and demo_file2_created):
             print("\nError: Failed to create one or both demo files. Aborting.")
             # Attempt cleanup of any file that *was* created
             if not args.keep_files:
                  if demo_file1_created and os.path.exists(demo_file1_name): os.remove(demo_file1_name)
                  if demo_file2_created and os.path.exists(demo_file2_name): os.remove(demo_file2_name)
             exit(1) # Exit if demo files couldn't be created

    else:
        print("\n--- Running in User-Specified File Mode ---")
        print(f"Using File 1: {input_file1}")
        print(f"Using File 2: {input_file2}")
        # Basic check if specified files exist
        if not os.path.exists(input_file1):
             print(f"Error: Specified File 1 not found: {input_file1}")
             exit(1)
        if not os.path.exists(input_file2):
             print(f"Error: Specified File 2 not found: {input_file2}")
             exit(1)


    # --- Call the End-to-End Comparison Function ---
    # Pass the determined input file paths and the keep_files flag
    results = compare_excel_files_e2e(input_file1, input_file2, keep_files=args.keep_files)

    # Optional: Print the returned results dictionary from the E2E function
    if results is not None:
         print("\n--- Comparison Results (Dictionary returned by E2E function) ---")
         # Use json.dumps for pretty printing the potentially large dictionary
         # Limit depth or handle large output appropriately if needed for console
         try:
             print(json.dumps(results, indent=2))
         except TypeError as e:
             print(f"Could not serialize results dictionary to JSON: {e}")
             print(results) # Print raw dictionary as fallback
    else:
         print("\nEnd-to-end comparison function did not return results (critical error occurred).")


    # --- Cleanup Demo Excel Files (if they were created and not keeping files) ---
    if use_demo_files and not args.keep_files:
        print("\n--- Cleaning Up Demo Excel Files ---")
        # Only clean up the specific demo files created by this run
        if demo_file1_created and os.path.exists(input_file1):
            try:
                os.remove(input_file1)
                print(f"Removed demo file: '{input_file1}'")
            except Exception as e:
                print(f"Error cleaning up demo file '{input_file1}': {e}")
        if demo_file2_created and os.path.exists(input_file2):
            try:
                os.remove(input_file2)
                print(f"Removed demo file: '{input_file2}'")
            except Exception as e:
                print(f"Error cleaning up demo file '{input_file2}': {e}")


    print("\nUtility finished.")

# --- Script Entry Point ---
if __name__ == "__main__":
    # Check for required libraries
    try:
        import openpyxl
    except ImportError:
        print("Error: The 'openpyxl' library is required but not installed.")
        print("Please install it using: pip install openpyxl")
        exit(1)
    try:
        import pytz
    except ImportError:
        print("Error: The 'pytz' library is required but not installed.")
        print("Please install it using: pip install pytz")
        exit(1)

    main()