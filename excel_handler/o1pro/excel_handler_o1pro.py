#!/usr/bin/env python3
"""
A single Python script demonstrating how to:

1) Create some dummy Excel files (for demo purposes).
2) Read each Excel file into a nested dictionary of the form:
   {
       "SheetName": {
           "A1": {
               "value": <evaluated_value>,
               "formula": <formula_string_if_any_or_empty>
           },
           "A2": { ... },
           ...
       },
       ...
   }
3) Compare two such dictionaries to see differences in cell values/formulas.
4) Print out / export the comparison result to text files, including:
   - A full dump of each Excel file's contents (file1_contents.txt & file2_contents.txt)
   - A comparison summary (comparison_summary.txt)

NEW REQUIREMENT:
 - By default, produce the text outputs in an "output_{datetime_Jakarta}" folder,
   e.g., "output_20230407_134500" for 2023-04-07 13:45:00 in Jakarta time.

All functionalities are kept in a single script, with dedicated functions to
achieve the tasks.
"""

import openpyxl
import os
import datetime
import pytz


def read_excel_to_dict(file_path):
    """
    Read an Excel file and return a nested dictionary of:
        {
            "SheetName": {
                "A1": {"value": <cell_value>, "formula": <cell_formula_or_empty>},
                "A2": {"value": <cell_value>, "formula": <cell_formula_or_empty>},
                ...
            },
            ...
        }

    We use two workbooks:
      - data_only=False: to extract the actual formulas.
      - data_only=True:  to extract the last computed (cached) values.

    Then we combine them to get both the cell's formula (if any) and
    its last computed value.
    """
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
    wb_values = openpyxl.load_workbook(file_path, data_only=True)

    excel_dict = {}

    # We'll iterate over wb_formulas' sheetnames because it definitely has
    # all the sheets in the file.
    for sheet_name in wb_formulas.sheetnames:
        sheet_formula = wb_formulas[sheet_name]
        sheet_value = wb_values[sheet_name]

        sheet_data = {}
        for row in sheet_formula.iter_rows():
            for cell_f in row:
                cell_v = sheet_value[cell_f.coordinate]

                # If the cell is a formula, openpyxl (with data_only=False)
                # stores that formula in 'cell_f.value' when cell_f.data_type == 'f'.
                if cell_f.data_type == 'f':
                    formula = cell_f.value  # e.g., "=SUM(A1,A2)"
                else:
                    formula = ""

                # cell_v.value is the computed value (from the data_only=True workbook),
                # or the literal value if there's no formula.
                computed_value = cell_v.value

                sheet_data[cell_f.coordinate] = {
                    "value": computed_value,
                    "formula": formula
                }

        excel_dict[sheet_name] = sheet_data

    return excel_dict


def compare_excel_dicts(dict1, dict2):
    """
    Compare two Excel dictionaries produced by `read_excel_to_dict`.
    Return a dictionary describing only the differences in cell contents for
    shared sheets.

    The returned dictionary structure is:
    {
       "Sheet1": {
          "A1": {
             "value": (val_in_dict1, val_in_dict2)     # only if they differ
             "formula": (formula_in_dict1, formula_in_dict2)  # only if they differ
          },
          "A2": {...},
          ...
       },
       "Sheet2": ...
    }

    Only cells that differ in "value" or "formula" will appear.
    Sheets that are not in both dict1 and dict2 won't appear here (this is a
    cell-level difference dictionary). We'll handle missing sheets separately.
    """
    differences = {}

    # Collect only the intersection of sheet names for cell-level diffs
    common_sheets = set(dict1.keys()).intersection(dict2.keys())

    for sheet in common_sheets:
        sheet_diff = {}
        sheet1_data = dict1[sheet]
        sheet2_data = dict2[sheet]

        # Collect all cell coords from both sheets in the intersection
        all_cells = set(sheet1_data.keys()).union(sheet2_data.keys())
        for cell_coord in all_cells:
            cell1 = sheet1_data.get(cell_coord, {})
            cell2 = sheet2_data.get(cell_coord, {})

            val1 = cell1.get("value", None)
            val2 = cell2.get("value", None)
            formula1 = cell1.get("formula", "")
            formula2 = cell2.get("formula", "")

            cell_subdiff = {}

            # Check value differences
            if val1 != val2:
                cell_subdiff["value"] = (val1, val2)

            # Check formula differences
            if formula1 != formula2:
                cell_subdiff["formula"] = (formula1, formula2)

            # If anything differs, record it
            if cell_subdiff:
                sheet_diff[cell_coord] = cell_subdiff

        if sheet_diff:
            differences[sheet] = sheet_diff

    return differences


def export_excel_dict_to_txt(excel_dict, file_path, source_excel_path=""):
    """
    Export the Excel dictionary to a text file for a full trace of contents.
    The format will be something like:

        Contents of {source_excel_path}:
        --------------------------------
        Sheet: Sheet1
          Cell A1 => value: 10, formula:
          Cell A2 => value: 20, formula:
          Cell A3 => value: 30, formula: =SUM(A1,A2)
        ...
        Sheet: Sheet2
          Cell C10 => value: 99, formula:
          Cell C11 => value: 198, formula: =C10*2
        ...

    Parameters:
    -----------
    excel_dict : dict
        The nested dictionary from read_excel_to_dict().
    file_path : str
        The path of the text file to output (including folder).
    source_excel_path : str (optional)
        The original Excel file path, just for reference in the header.
    """
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(f"Contents of {source_excel_path}:\n")
        f.write("-" * 40 + "\n")

        for sheet_name in excel_dict:
            f.write(f"Sheet: {sheet_name}\n")
            sheet_data = excel_dict[sheet_name]
            # Sort the cell coordinates for a neat listing (e.g., A1, A2, A3, ...)
            for cell_coord in sorted(sheet_data.keys()):
                cell_info = sheet_data[cell_coord]
                value_str = str(cell_info["value"])
                formula_str = cell_info["formula"] if cell_info["formula"] else ""
                f.write(f"  Cell {cell_coord} => value: {value_str}, formula: {formula_str}\n")
            f.write("\n")  # extra space between sheets


def export_comparison_to_txt(diff_dict, dict1, dict2, file_path,
                             excel1_name="", excel2_name=""):
    """
    Export the comparison summary to a text file. This includes:

    - Which sheets exist in one file but not the other
    - Sheets that exist in both
    - Cell-level differences in the common sheets

    Parameters:
    -----------
    diff_dict : dict
        The differences dictionary returned by compare_excel_dicts().
    dict1, dict2 : dict
        The dictionaries returned by read_excel_to_dict() for each file.
    file_path : str
        The path of the text file to output (including folder).
    excel1_name, excel2_name : str
        The original Excel file paths, just for reference in the header.
    """
    sheets1 = set(dict1.keys())
    sheets2 = set(dict2.keys())

    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(f"Comparison summary between {excel1_name} and {excel2_name}\n")
        f.write("=" * 60 + "\n\n")

        # 1) Compare sheet sets
        only_in_1 = sheets1 - sheets2
        only_in_2 = sheets2 - sheets1
        in_both = sheets1 & sheets2

        f.write("Sheet differences:\n")
        if only_in_1:
            f.write(f"  Sheets only in {excel1_name}: {', '.join(only_in_1)}\n")
        if only_in_2:
            f.write(f"  Sheets only in {excel2_name}: {', '.join(only_in_2)}\n")
        if not only_in_1 and not only_in_2:
            f.write("  No sheet-level differences found (same sheet names).\n")
        f.write("\n")

        # 2) Cell-level differences in sheets that appear in both
        f.write("Cell-level differences in common sheets:\n")
        if not diff_dict:
            f.write("  No cell-level differences found.\n")
        else:
            for sheet_name in diff_dict:
                f.write(f"  Differences in sheet '{sheet_name}':\n")
                cell_diffs = diff_dict[sheet_name]
                for cell_coord, changes in cell_diffs.items():
                    f.write(f"    Cell {cell_coord}:\n")
                    for attr, (val1, val2) in changes.items():
                        f.write(f"      {attr} differs: '{val1}' vs. '{val2}'\n")
                f.write("\n")

        f.write("End of comparison.\n")


def create_dummy_excel_files():
    """
    Create two dummy Excel files to demonstrate usage:
      - file1.xlsx
      - file2.xlsx

    They will have:
      - Some sheets in both files with slight differences
      - At least one sheet that appears only in file1
      - At least one sheet that appears only in file2
    """
    # --------------- File 1 ---------------
    wb1 = openpyxl.Workbook()
    ws1_1 = wb1.active  # First sheet
    ws1_1.title = "Sheet1"

    # Populate some cells
    ws1_1["A1"] = 10
    ws1_1["A2"] = 20
    ws1_1["A3"] = "=SUM(A1, A2)"  # Formula

    ws1_1["B1"] = "Hello"
    ws1_1["B2"] = "=CONCAT(B1, \" World\")"  # Another formula

    # Create a second sheet (shared name with file2, but data might differ)
    ws1_2 = wb1.create_sheet("Sheet2")
    ws1_2["C10"] = 99
    ws1_2["C11"] = "=C10 * 2"

    # Create a third sheet that does NOT appear in file2
    ws1_3 = wb1.create_sheet("SheetX_OnlyInFile1")
    ws1_3["A1"] = 123
    ws1_3["A2"] = "=A1 * 2"

    wb1.save("file1.xlsx")

    # --------------- File 2 ---------------
    wb2 = openpyxl.Workbook()
    ws2_1 = wb2.active
    ws2_1.title = "Sheet1"

    # Slightly different data
    ws2_1["A1"] = 10
    ws2_1["A2"] = 21   # changed
    ws2_1["A3"] = "=SUM(A1, A2)"  # formula is the same

    ws2_1["B1"] = "Hello"
    ws2_1["B2"] = "=CONCAT(B1, \" Universe\")"  # changed from "World" to "Universe"

    # Keep "Sheet2" to test "common sheet" differences, but slightly different values
    ws2_2 = wb2.create_sheet("Sheet2")
    ws2_2["C10"] = 99
    ws2_2["C11"] = "=C10 * 3"  # changed multiplier from 2 -> 3

    # A third sheet that does not appear in file1
    ws2_3 = wb2.create_sheet("UniqueSheetInFile2")
    ws2_3["A1"] = "Only in File2"

    wb2.save("file2.xlsx")


def main():
    # 1) Create dummy Excel files for demonstration:
    create_dummy_excel_files()
    print("[INFO] Created 'file1.xlsx' and 'file2.xlsx' for demonstration.\n")

    # 2) Read each file into dictionaries
    dict1 = read_excel_to_dict("file1.xlsx")
    dict2 = read_excel_to_dict("file2.xlsx")

    # 3) Create an output folder named "output_{datetime_Jakarta}"
    #    Example: output_20250407_101530
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    now_jakarta = datetime.datetime.now(tz=jakarta_tz)
    folder_name = now_jakarta.strftime("output_%Y%m%d_%H%M%S")
    os.makedirs(folder_name, exist_ok=True)
    print(f"[INFO] Outputs will be written to folder: {folder_name}\n")

    # 4) Export each dictionary to its own text file in the new folder
    file1_out_path = os.path.join(folder_name, "file1_contents.txt")
    file2_out_path = os.path.join(folder_name, "file2_contents.txt")

    export_excel_dict_to_txt(dict1, file1_out_path, "file1.xlsx")
    export_excel_dict_to_txt(dict2, file2_out_path, "file2.xlsx")
    print("[INFO] Exported the contents of file1.xlsx and file2.xlsx to text files.\n")

    # 5) Compare the two dictionaries at cell level (only in shared sheets)
    diff = compare_excel_dicts(dict1, dict2)

    # 6) Export comparison summary (sheet-level & cell-level differences) to a text file
    comparison_out_path = os.path.join(folder_name, "comparison_summary.txt")
    export_comparison_to_txt(diff, dict1, dict2, comparison_out_path, "file1.xlsx", "file2.xlsx")
    print("[INFO] Exported comparison summary to 'comparison_summary.txt'.\n")

    # (Optional) If you want to keep the Excel files, leave them as is.
    # If you want to clean up, uncomment these lines:
    #
    # os.remove("file1.xlsx")
    # os.remove("file2.xlsx")
    #
    print("[INFO] Done. Check your output folder and text files.")


if __name__ == "__main__":
    main()
