import openpyxl
import os
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- Core Data Extraction Functions ---

def get_cell_data(cell_value, cell_formula):
    """
    Processes a cell's value and formula.

    Args:
        cell_value: The evaluated value of the cell (obtained using data_only=True).
        cell_formula: The formula string if the cell contains one, otherwise potentially
                      the static value (obtained using data_only=False).

    Returns:
        A dictionary containing the 'value' and 'formula'.
        Returns {'value': cell_value, 'formula': ""} if the cell doesn't contain a formula.
        Returns {'value': cell_value, 'formula': cell_formula} if it has a formula.
    """
    formula_str = ""
    # Check if the content from the formula-aware load looks like a formula
    if isinstance(cell_formula, str) and cell_formula.startswith('='):
        formula_str = cell_formula
    # Sometimes numeric/boolean values might be cached results of formulas,
    # but openpyxl might not store the '=' if the type isn't 'f'.
    # Relying on the data_only=False load giving the formula string is safer.

    return {"value": cell_value, "formula": formula_str}

def read_worksheet_data(ws_values, ws_formulas):
    """
    Reads all cell data (value and formula) from a single worksheet.

    Args:
        ws_values: An openpyxl worksheet object loaded with data_only=True.
        ws_formulas: An openpyxl worksheet object loaded with data_only=False.

    Returns:
        A dictionary where keys are cell coordinates (e.g., "A1") and
        values are dictionaries {'value': ..., 'formula': ...}.
    """
    sheet_data = {}
    # Iterate through rows/columns based on the formula sheet's dimensions,
    # as it preserves formula strings. Assume dimensions are the same,
    # which is generally true when loading the same sheet twice.
    for row_idx in range(1, ws_formulas.max_row + 1):
        for col_idx in range(1, ws_formulas.max_column + 1):
            cell_coord = f"{get_column_letter(col_idx)}{row_idx}"

            # Get the formula string (or static value) from the formula sheet
            cell_formulas_obj = ws_formulas.cell(row=row_idx, column=col_idx)
            formula_content = cell_formulas_obj.value # This is the key part for getting formulas

            # Get the evaluated value from the value sheet
            cell_values_obj = ws_values.cell(row=row_idx, column=col_idx)
            evaluated_value = cell_values_obj.value

            # Only store data if the cell isn't completely empty in both views
            # (or adjust this condition if truly empty cells need tracking)
            # A cell might have a formula resulting in "" or 0, which we want to keep.
            # Let's store all cells within the max range for comparison consistency.
            # if evaluated_value is not None or (isinstance(formula_content, str) and formula_content.startswith('=')):

            cell_info = get_cell_data(evaluated_value, formula_content)
            sheet_data[cell_coord] = cell_info

    return sheet_data

def read_excel_file_data(filepath):
    """
    Reads all data (values and formulas) from all worksheets in an Excel file.

    Args:
        filepath (str): The path to the Excel file (.xlsx).

    Returns:
        A nested dictionary representing the Excel file's data:
        {
            "sheet_name_1": {
                "A1": {"value": ..., "formula": ...},
                "B2": {"value": ..., "formula": ...},
                ...
            },
            "sheet_name_2": { ... },
            ...
        }
        Returns None if the file cannot be read.
    """
    try:
        # Load workbook once to get evaluated values (formulas computed)
        wb_values = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        # Load workbook again to get formulas as strings
        wb_formulas = openpyxl.load_workbook(filepath, data_only=False, read_only=True)

    except FileNotFoundError:
        print(f"Error: File not found at '{filepath}'")
        return None
    except Exception as e:
        print(f"Error reading Excel file '{filepath}': {e}")
        return None

    all_excel_data = {}
    # Use sheet names from the formula workbook as the primary source
    for sheet_name in wb_formulas.sheetnames:
        # Ensure the sheet exists in both loaded workbooks (should normally be true)
        if sheet_name in wb_values.sheetnames:
            ws_values = wb_values[sheet_name]
            ws_formulas = wb_formulas[sheet_name]
            print(f"  Processing sheet: '{sheet_name}'...")
            all_excel_data[sheet_name] = read_worksheet_data(ws_values, ws_formulas)
        else:
             print(f"  Warning: Sheet '{sheet_name}' found in formula view but not value view. Skipping.")

    # Close workbooks if opened in read-only mode (good practice, though less critical)
    # Openpyxl's read-only mode often handles this implicitly upon exit/garbage collection.
    # wb_values.close()
    # wb_formulas.close()

    return all_excel_data

# --- Comparison Function ---

def compare_excel_data(data1, data2):
    """
    Compares the extracted data from two Excel files.

    Args:
        data1 (dict): The nested dictionary data from the first Excel file.
        data2 (dict): The nested dictionary data from the second Excel file.

    Returns:
        A dictionary highlighting the differences:
        {
            "sheet_name": {
                "cell_coord": {
                    "file1": {"value": v1, "formula": f1},
                    "file2": {"value": v2, "formula": f2}
                },
                ... # Other differing cells
            },
            ... # Other sheets with differences
            "_metadata": { # Information about structure differences
                 "sheets_only_in_file1": [...],
                 "sheets_only_in_file2": [...]
            }
        }
        Returns an empty dictionary if no differences are found.
    """
    if data1 is None or data2 is None:
        print("Error: Cannot compare None data.")
        return {"_metadata": {"error": "Input data missing"}}

    differences = {}
    all_sheets = set(data1.keys()) | set(data2.keys())
    sheets_only_in_file1 = list(set(data1.keys()) - set(data2.keys()))
    sheets_only_in_file2 = list(set(data2.keys()) - set(data1.keys()))

    meta = {
        "sheets_only_in_file1": sorted(sheets_only_in_file1),
        "sheets_only_in_file2": sorted(sheets_only_in_file2)
    }
    if meta["sheets_only_in_file1"] or meta["sheets_only_in_file2"]:
        differences["_metadata"] = meta

    common_sheets = set(data1.keys()) & set(data2.keys())

    for sheet_name in common_sheets:
        sheet_diff = {}
        sheet_data1 = data1.get(sheet_name, {})
        sheet_data2 = data2.get(sheet_name, {})
        all_cells = set(sheet_data1.keys()) | set(sheet_data2.keys())

        for cell_coord in sorted(list(all_cells)): # Sort for consistent output
            cell_info1 = sheet_data1.get(cell_coord)
            cell_info2 = sheet_data2.get(cell_coord)

            # Define defaults for comparison if a cell exists in one sheet but not the other
            default_cell = {"value": None, "formula": ""}
            cell_info1 = cell_info1 if cell_info1 is not None else default_cell
            cell_info2 = cell_info2 if cell_info2 is not None else default_cell

            # Compare value and formula
            values_differ = cell_info1["value"] != cell_info2["value"]
            formulas_differ = cell_info1["formula"] != cell_info2["formula"]

            # Note: This comparison is strict (e.g., 1.0 != 1).
            # You might want to add tolerance for floating-point comparisons if needed.

            if values_differ or formulas_differ:
                sheet_diff[cell_coord] = {
                    "file1": cell_info1,
                    "file2": cell_info2
                }

        if sheet_diff:
            differences[sheet_name] = sheet_diff

    return differences

# --- Demo File Creation ---

def create_demo_excel_file(filepath, file_data):
    """Creates a demo .xlsx file with specified data."""
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for sheet_name, sheet_contents in file_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for cell_coord, content in sheet_contents.items():
            # Content can be a direct value or a dictionary {'value': ..., 'formula': ...}
            # For simplicity here, we just write the 'value' which could be a formula string
            if isinstance(content, dict) and 'value' in content:
                 # If the value itself looks like a formula, write it directly
                 # Otherwise, just write the value. openpyxl handles types.
                 ws[cell_coord] = content['value']
            else: # Assume direct value if not a dict
                 ws[cell_coord] = content

            # If a formula was explicitly provided in the dict, write it
            # Note: openpyxl writes formulas just by assigning a string starting with '='
            if isinstance(content, dict) and 'formula' in content and content['formula']:
                 ws[cell_coord] = content['formula'] # Overwrites previous if needed


    try:
        wb.save(filepath)
        print(f"Successfully created demo file: '{filepath}'")
    except Exception as e:
        print(f"Error creating demo file '{filepath}': {e}")
    finally:
        wb.close()


# --- Main Execution & Demonstration ---

def main():
    """
    Main function to demonstrate reading and comparing Excel files.
    """
    print("Starting Excel Comparison Utility Demo...")

    # Define Demo Data
    file1_name = "demo_excel_1.xlsx"
    file2_name = "demo_excel_2.xlsx"

    file1_data_structure = {
        "DataSheet": {
            "A1": {"value": "Name", "formula": ""},
            "B1": {"value": "Value1", "formula": ""},
            "C1": {"value": "Value2", "formula": ""},
            "D1": {"value": "Total", "formula": ""},
            "A2": {"value": "Item A", "formula": ""},
            "B2": {"value": 10, "formula": ""},
            "C2": {"value": 20, "formula": ""},
            "D2": {"value": 30, "formula": "=B2+C2"}, # Formula
            "A3": {"value": "Item B", "formula": ""},
            "B3": {"value": 15, "formula": ""},
            "C3": {"value": 25, "formula": ""},
            "D3": {"value": 40, "formula": "=SUM(B3:C3)"}, # Different Formula type
        },
        "Summary": {
            "A1": {"value": "Grand Total", "formula": ""},
            "B1": {"value": 70, "formula": "=DataSheet!D2+DataSheet!D3"},
             "C5": {"value": "Only in File 1 Sheet", "formula": ""}
        }
    }

    file2_data_structure = {
        "DataSheet": { # Same sheet name
            "A1": {"value": "Name", "formula": ""}, # Same
            "B1": {"value": "Value One", "formula": ""}, # Different Value
            "C1": {"value": "Value2", "formula": ""}, # Same
            "D1": {"value": "Total Sum", "formula": ""}, # Different Value
            "A2": {"value": "Item A", "formula": ""}, # Same
            "B2": {"value": 10, "formula": ""}, # Same
            "C2": {"value": 20, "formula": ""}, # Same
            "D2": {"value": 30, "formula": "=B2 + C2"}, # Same Formula (but space added maybe?) -> check if comparison catches it
            "A3": {"value": "Item B", "formula": ""}, # Same
            "B3": {"value": 15, "formula": ""}, # Same
            "C3": {"value": 99, "formula": ""}, # Different Value
            "D3": {"value": 114, "formula": "=B3+C3"}, # Different Formula (SUM vs +) and different value
            "E5": {"value": "Extra Col Cell", "formula": ""} # Extra Cell
        },
        "Details": { # Different sheet name
            "X1": {"value": "Detail Info", "formula": ""}
        },
         "Summary": { # Same sheet name, different content
            "A1": {"value": "Overall Total", "formula": ""}, # Diff Value
            "B1": {"value": 144, "formula": "=DataSheet!D2+DataSheet!D3+DataSheet!C3"}, # Diff Value & Formula
             "Z9": {"value": "Only in File 2 Sheet", "formula": ""}
        }
    }

    # 1. Create Demo Files
    print("\n--- Creating Demo Files ---")
    create_demo_excel_file(file1_name, file1_data_structure)
    create_demo_excel_file(file2_name, file2_data_structure)

    # 2. Read Data from Files
    print(f"\n--- Reading Data from '{file1_name}' ---")
    excel1_contents = read_excel_file_data(file1_name)

    print(f"\n--- Reading Data from '{file2_name}' ---")
    excel2_contents = read_excel_file_data(file2_name)

    # Optional: Print extracted data (can be large)
    # print("\n--- Extracted Data File 1 ---")
    # if excel1_contents:
    #     print(json.dumps(excel1_contents, indent=2))
    # else:
    #     print("Failed to read data from file 1.")

    # print("\n--- Extracted Data File 2 ---")
    # if excel2_contents:
    #     print(json.dumps(excel2_contents, indent=2))
    # else:
    #     print("Failed to read data from file 2.")

    # 3. Compare the Data
    print("\n--- Comparing Files ---")
    comparison_results = compare_excel_data(excel1_contents, excel2_contents)

    # 4. Print Comparison Results
    print("\n--- Comparison Results ---")
    if not comparison_results:
        print("No differences found between the files.")
    else:
        print(json.dumps(comparison_results, indent=2))

        # Provide a summary of findings
        print("\n--- Summary of Differences ---")
        meta = comparison_results.get("_metadata", {})
        if meta.get("sheets_only_in_file1"):
            print(f"Sheets only in {file1_name}: {meta['sheets_only_in_file1']}")
        if meta.get("sheets_only_in_file2"):
            print(f"Sheets only in {file2_name}: {meta['sheets_only_in_file2']}")

        for sheet, diffs in comparison_results.items():
            if sheet == "_metadata": continue
            print(f"\nDifferences found in sheet: '{sheet}'")
            cell_count = len(diffs)
            print(f"  Number of differing cells: {cell_count}")
            # Example of listing a few differing cells:
            limit = 5
            count = 0
            for cell, details in diffs.items():
                 if count < limit:
                     print(f"  - Cell '{cell}':")
                     print(f"    {file1_name}: Value='{details['file1']['value']}', Formula='{details['file1']['formula']}'")
                     print(f"    {file2_name}: Value='{details['file2']['value']}', Formula='{details['file2']['formula']}'")
                     count += 1
                 else:
                     print(f"    ... and {cell_count - limit} more differing cells in this sheet.")
                     break


    # 5. Cleanup Demo Files
    print("\n--- Cleaning Up Demo Files ---")
    try:
        if os.path.exists(file1_name):
            os.remove(file1_name)
            print(f"Removed '{file1_name}'")
        if os.path.exists(file2_name):
            os.remove(file2_name)
            print(f"Removed '{file2_name}'")
    except Exception as e:
        print(f"Error during cleanup: {e}")

    print("\nDemo finished.")

# --- Script Entry Point ---
if __name__ == "__main__":
    # Check for required library
    try:
        import openpyxl
    except ImportError:
        print("Error: The 'openpyxl' library is required but not installed.")
        print("Please install it using: pip install openpyxl")
        exit(1) # Exit if library is missing

    main()